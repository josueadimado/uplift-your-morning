"""
Custom admin management views for staff users.
These views allow staff to manage content without needing full Django admin access.
"""
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View
from django.conf import settings
import requests
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.db.models import Sum, Q, Count
from django.http import HttpResponse, HttpResponseBadRequest
from django.template.loader import render_to_string
import csv
import io
from datetime import datetime

from apps.devotions.models import Devotion, DevotionSeries
from django.http import JsonResponse
import json
from apps.events.models import Event
from apps.resources.models import Resource, ResourceCategory, FortyDaysNote, FortyDaysNoteCategory
from apps.community.models import PrayerRequest, Testimony
from apps.pages.models import Donation, FortyDaysConfig, SiteSettings, CounselingBooking, Pledge
from apps.subscriptions.models import Subscriber, ScheduledNotification


class StaffRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Mixin to ensure only staff users can access these views."""
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        """
        Add a flag so the base template knows we're in the admin area.
        This lets us hide the public top navigation/footer.
        """
        context = super().get_context_data(**kwargs)
        context['hide_main_nav'] = True
        return context


# ==================== DEVOTIONS ====================

class DevotionListView(StaffRequiredMixin, ListView):
    """List all devotions with filters."""
    model = Devotion
    template_name = 'admin/devotions/list.html'
    context_object_name = 'devotions'
    paginate_by = 20

    def get_queryset(self):
        queryset = Devotion.objects.all()
        # Filter by published status
        status = self.request.GET.get('status')
        if status == 'published':
            queryset = queryset.filter(is_published=True)
        elif status == 'draft':
            queryset = queryset.filter(is_published=False)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(body__icontains=search) |
                Q(scripture_reference__icontains=search)
            )
        return queryset.order_by('-publish_date', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = Devotion.objects.count()
        context['published_count'] = Devotion.objects.filter(is_published=True).count()
        context['draft_count'] = Devotion.objects.filter(is_published=False).count()
        return context


class DevotionCreateView(StaffRequiredMixin, CreateView):
    """Create a new devotion."""
    model = Devotion
    template_name = 'admin/devotions/form.html'
    fields = [
        'title', 'series', 'theme', 'scripture_reference', 'passage_text',
        'body', 'quote', 'reflection', 'prayer', 'action_point', 'publish_date',
        'is_published', 'image', 'audio_file', 'pdf_file', 'featured'
    ]
    success_url = reverse_lazy('manage:devotions_list')

    def get_initial(self):
        """
        Prefill publish date to today and default to published.
        """
        from django.utils import timezone
        initial = super().get_initial()
        initial.update({
            'publish_date': timezone.localdate(),
            'is_published': True,
        })
        return initial

    def get_context_data(self, **kwargs):
        """Add day count to context."""
        context = super().get_context_data(**kwargs)
        # Calculate day count for initial date (today)
        from django.utils import timezone
        publish_date = self.get_initial().get('publish_date', timezone.localdate())
        if publish_date:
            day_of_year = publish_date.timetuple().tm_yday
            year = publish_date.year
            context['day_count'] = f"Day {day_of_year} of year {year}"
        else:
            context['day_count'] = ""
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Devotion created successfully!')
        return super().form_valid(form)


class DevotionUpdateView(StaffRequiredMixin, UpdateView):
    """Edit an existing devotion."""
    model = Devotion
    template_name = 'admin/devotions/form.html'
    fields = [
        'title', 'series', 'theme', 'scripture_reference', 'passage_text',
        'body', 'quote', 'reflection', 'prayer', 'action_point', 'publish_date',
        'is_published', 'image', 'audio_file', 'pdf_file', 'featured'
    ]
    success_url = reverse_lazy('manage:devotions_list')

    def get_context_data(self, **kwargs):
        """Add day count to context."""
        context = super().get_context_data(**kwargs)
        # Calculate day count from the devotion's publish date
        if self.object and self.object.publish_date:
            context['day_count'] = self.object.get_day_count()
        else:
            # If no date yet, calculate from form initial or today
            publish_date = self.object.publish_date if self.object else None
            if not publish_date:
                from django.utils import timezone
                publish_date = timezone.localdate()
            if publish_date:
                day_of_year = publish_date.timetuple().tm_yday
                year = publish_date.year
                context['day_count'] = f"Day {day_of_year} of year {year}"
            else:
                context['day_count'] = ""
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Devotion updated successfully!')
        return super().form_valid(form)


class DevotionDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a devotion."""
    model = Devotion
    template_name = 'admin/devotions/delete.html'
    success_url = reverse_lazy('manage:devotions_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Devotion deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== EVENTS ====================

class EventListView(StaffRequiredMixin, ListView):
    """List all events."""
    model = Event
    template_name = 'admin/events/list.html'
    context_object_name = 'events'
    paginate_by = 20

    def get_queryset(self):
        queryset = Event.objects.all()
        # Filter by upcoming/past
        filter_type = self.request.GET.get('filter')
        now = timezone.now()
        if filter_type == 'upcoming':
            queryset = queryset.filter(start_datetime__gte=now)
        elif filter_type == 'past':
            queryset = queryset.filter(start_datetime__lt=now)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(location__icontains=search)
            )
        return queryset.order_by('-start_datetime')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        context['total'] = Event.objects.count()
        context['upcoming_count'] = Event.objects.filter(start_datetime__gte=now).count()
        context['past_count'] = Event.objects.filter(start_datetime__lt=now).count()
        return context


class EventCreateView(StaffRequiredMixin, CreateView):
    """Create a new event."""
    model = Event
    template_name = 'admin/events/form.html'
    fields = [
        'title', 'description', 'start_datetime', 'end_datetime', 'location',
        'is_online', 'poster_image', 'livestream_url', 'youtube_url',
        'facebook_url', 'zoom_url', 'registration_open'
    ]
    success_url = reverse_lazy('manage:events_list')

    def form_valid(self, form):
        messages.success(self.request, 'Event created successfully!')
        return super().form_valid(form)


class EventUpdateView(StaffRequiredMixin, UpdateView):
    """Edit an existing event."""
    model = Event
    template_name = 'admin/events/form.html'
    fields = [
        'title', 'description', 'start_datetime', 'end_datetime', 'location',
        'is_online', 'poster_image', 'livestream_url', 'youtube_url',
        'facebook_url', 'zoom_url', 'registration_open'
    ]
    success_url = reverse_lazy('manage:events_list')

    def form_valid(self, form):
        messages.success(self.request, 'Event updated successfully!')
        return super().form_valid(form)


class EventDeleteView(StaffRequiredMixin, DeleteView):
    """Delete an event."""
    model = Event
    template_name = 'admin/events/delete.html'
    success_url = reverse_lazy('manage:events_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Event deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== RESOURCES ====================

class ResourceListView(StaffRequiredMixin, ListView):
    """List all resources."""
    model = Resource
    template_name = 'admin/resources/list.html'
    context_object_name = 'resources'
    paginate_by = 20

    def get_queryset(self):
        queryset = Resource.objects.all()
        # Filter by type
        resource_type = self.request.GET.get('type')
        if resource_type:
            queryset = queryset.filter(type=resource_type)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = Resource.objects.count()
        return context


class ResourceCreateView(StaffRequiredMixin, CreateView):
    """Create a new resource."""
    model = Resource
    template_name = 'admin/resources/form.html'
    fields = ['title', 'category', 'type', 'description', 'file', 'external_url', 'is_featured']
    success_url = reverse_lazy('manage:resources_list')

    def form_valid(self, form):
        messages.success(self.request, 'Resource created successfully!')
        return super().form_valid(form)


class ResourceUpdateView(StaffRequiredMixin, UpdateView):
    """Edit an existing resource."""
    model = Resource
    template_name = 'admin/resources/form.html'
    fields = ['title', 'category', 'type', 'description', 'file', 'external_url', 'is_featured']
    success_url = reverse_lazy('manage:resources_list')

    def form_valid(self, form):
        messages.success(self.request, 'Resource updated successfully!')
        return super().form_valid(form)


class ResourceDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a resource."""
    model = Resource
    template_name = 'admin/resources/delete.html'
    success_url = reverse_lazy('manage:resources_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Resource deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== RESOURCE CATEGORIES ====================

class ResourceCategoryListView(StaffRequiredMixin, ListView):
    """List all resource categories."""
    model = ResourceCategory
    template_name = 'admin/resources/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20

    def get_queryset(self):
        queryset = ResourceCategory.objects.all()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = ResourceCategory.objects.count()
        return context


class ResourceCategoryCreateView(StaffRequiredMixin, CreateView):
    """Create a new resource category."""
    model = ResourceCategory
    template_name = 'admin/resources/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('manage:resource_categories_list')

    def form_valid(self, form):
        messages.success(self.request, 'Resource category created successfully!')
        return super().form_valid(form)


class ResourceCategoryUpdateView(StaffRequiredMixin, UpdateView):
    """Update a resource category."""
    model = ResourceCategory
    template_name = 'admin/resources/category_form.html'
    fields = ['name', 'description']
    success_url = reverse_lazy('manage:resource_categories_list')

    def form_valid(self, form):
        messages.success(self.request, 'Resource category updated successfully!')
        return super().form_valid(form)


class ResourceCategoryDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a resource category."""
    model = ResourceCategory
    template_name = 'admin/resources/category_delete.html'
    success_url = reverse_lazy('manage:resource_categories_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Resource category deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== 40 DAYS NOTES ====================

class FortyDaysNoteCategoryListView(StaffRequiredMixin, ListView):
    """List all 40 Days note categories."""
    model = FortyDaysNoteCategory
    template_name = 'admin/fortydays/note_category_list.html'
    context_object_name = 'categories'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = FortyDaysNoteCategory.objects.count()
        return context


class FortyDaysNoteCategoryCreateView(StaffRequiredMixin, CreateView):
    """Create a new 40 Days note category."""
    model = FortyDaysNoteCategory
    template_name = 'admin/fortydays/note_category_form.html'
    fields = ['name', 'description', 'order']
    success_url = reverse_lazy('manage:fortydays_note_categories_list')

    def form_valid(self, form):
        messages.success(self.request, '40 Days note category created successfully!')
        return super().form_valid(form)


class FortyDaysNoteCategoryUpdateView(StaffRequiredMixin, UpdateView):
    """Update a 40 Days note category."""
    model = FortyDaysNoteCategory
    template_name = 'admin/fortydays/note_category_form.html'
    fields = ['name', 'description', 'order']
    success_url = reverse_lazy('manage:fortydays_note_categories_list')

    def form_valid(self, form):
        messages.success(self.request, '40 Days note category updated successfully!')
        return super().form_valid(form)


class FortyDaysNoteCategoryDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a 40 Days note category."""
    model = FortyDaysNoteCategory
    template_name = 'admin/fortydays/note_category_delete.html'
    success_url = reverse_lazy('manage:fortydays_note_categories_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, '40 Days note category deleted successfully!')
        return super().delete(request, *args, **kwargs)


class FortyDaysNoteListView(StaffRequiredMixin, ListView):
    """List all 40 Days notes."""
    model = FortyDaysNote
    template_name = 'admin/fortydays/note_list.html'
    context_object_name = 'notes'
    paginate_by = 20

    def get_queryset(self):
        queryset = FortyDaysNote.objects.all()
        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        # Filter by published status
        status = self.request.GET.get('status')
        if status == 'published':
            queryset = queryset.filter(is_published=True)
        elif status == 'unpublished':
            queryset = queryset.filter(is_published=False)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                title__icontains=search
            ) | queryset.filter(
                content__icontains=search
            ) | queryset.filter(
                expert_name__icontains=search
            )
        return queryset.order_by('-session_date', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = FortyDaysNoteCategory.objects.all()
        context['total'] = FortyDaysNote.objects.count()
        context['published_count'] = FortyDaysNote.objects.filter(is_published=True).count()
        context['unpublished_count'] = FortyDaysNote.objects.filter(is_published=False).count()
        context['current_category'] = self.request.GET.get('category')
        context['current_status'] = self.request.GET.get('status')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class FortyDaysNoteCreateView(StaffRequiredMixin, CreateView):
    """Create a new 40 Days note."""
    model = FortyDaysNote
    template_name = 'admin/fortydays/note_form.html'
    fields = [
        'title', 'category', 'banner_image', 'content', 'expert_name',
        'session_date', 'is_published', 'is_featured'
    ]
    success_url = reverse_lazy('manage:fortydays_notes_list')

    def form_valid(self, form):
        messages.success(self.request, '40 Days note created successfully!')
        return super().form_valid(form)


class FortyDaysNoteUpdateView(StaffRequiredMixin, UpdateView):
    """Update a 40 Days note."""
    model = FortyDaysNote
    template_name = 'admin/fortydays/note_form.html'
    fields = [
        'title', 'category', 'banner_image', 'content', 'expert_name',
        'session_date', 'is_published', 'is_featured'
    ]
    success_url = reverse_lazy('manage:fortydays_notes_list')

    def form_valid(self, form):
        messages.success(self.request, '40 Days note updated successfully!')
        return super().form_valid(form)


class FortyDaysNoteDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a 40 Days note."""
    model = FortyDaysNote
    template_name = 'admin/fortydays/note_delete.html'
    success_url = reverse_lazy('manage:fortydays_notes_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, '40 Days note deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== PRAYER REQUESTS ====================

class PrayerRequestListView(StaffRequiredMixin, ListView):
    """List all prayer requests."""
    model = PrayerRequest
    template_name = 'admin/prayers/list.html'
    context_object_name = 'prayers'
    paginate_by = 20

    def get_queryset(self):
        queryset = PrayerRequest.objects.all()
        # Filter by status
        status = self.request.GET.get('status')
        if status == 'open':
            queryset = queryset.filter(is_prayed_for=False)
        elif status == 'prayed':
            queryset = queryset.filter(is_prayed_for=True)
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = PrayerRequest.objects.count()
        context['open_count'] = PrayerRequest.objects.filter(is_prayed_for=False).count()
        context['prayed_count'] = PrayerRequest.objects.filter(is_prayed_for=True).count()
        return context


class PrayerRequestMarkPrayedView(StaffRequiredMixin, View):
    """Mark a prayer request as prayed for."""
    def post(self, request, *args, **kwargs):
        try:
            prayer = PrayerRequest.objects.get(pk=kwargs['pk'])
            prayer.is_prayed_for = True
            prayer.save()
            messages.success(request, 'Prayer request marked as prayed for!')
        except PrayerRequest.DoesNotExist:
            messages.error(request, 'Prayer request not found.')
        return redirect('manage:prayers_list')


class PrayerRequestDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a prayer request."""
    model = PrayerRequest
    template_name = 'admin/prayers/delete.html'
    success_url = reverse_lazy('manage:prayers_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Prayer request deleted successfully!')
        return super().delete(request, *args, **kwargs)


class PrayerRequestExportCSVView(StaffRequiredMixin, View):
    """Export prayer requests as CSV."""
    def get(self, request, *args, **kwargs):
        queryset = self._get_queryset()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="prayer_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Email', 'Request', 'Is Public', 'Is Prayed For', 'Date Submitted'])
        
        for prayer in queryset:
            writer.writerow([
                prayer.name or 'Anonymous',
                prayer.email or '',
                prayer.request,
                'Yes' if prayer.is_public else 'No',
                'Yes' if prayer.is_prayed_for else 'No',
                prayer.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = PrayerRequest.objects.all()
        status = self.request.GET.get('status')
        if status == 'open':
            queryset = queryset.filter(is_prayed_for=False)
        elif status == 'prayed':
            queryset = queryset.filter(is_prayed_for=True)
        return queryset.order_by('-created_at')


class PrayerRequestExportExcelView(StaffRequiredMixin, View):
    """Export prayer requests as Excel."""
    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponseBadRequest('Excel export requires openpyxl. Please install it: pip install openpyxl')
        
        queryset = self._get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Prayer Requests"
        
        # Header row
        headers = ['Name', 'Email', 'Request', 'Is Public', 'Is Prayed For', 'Date Submitted']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, prayer in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=prayer.name or 'Anonymous')
            ws.cell(row=row_num, column=2, value=prayer.email or '')
            ws.cell(row=row_num, column=3, value=prayer.request)
            ws.cell(row=row_num, column=4, value='Yes' if prayer.is_public else 'No')
            ws.cell(row=row_num, column=5, value='Yes' if prayer.is_prayed_for else 'No')
            ws.cell(row=row_num, column=6, value=prayer.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="prayer_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = PrayerRequest.objects.all()
        status = self.request.GET.get('status')
        if status == 'open':
            queryset = queryset.filter(is_prayed_for=False)
        elif status == 'prayed':
            queryset = queryset.filter(is_prayed_for=True)
        return queryset.order_by('-created_at')


class PrayerRequestExportPDFView(StaffRequiredMixin, View):
    """Export prayer requests as PDF (spreadsheet format)."""
    def get(self, request, *args, **kwargs):
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            return HttpResponseBadRequest('PDF export requires reportlab. Please install it: pip install reportlab')
        
        queryset = self._get_queryset()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        story = []
        story.append(Paragraph("Prayer Requests Export", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Prepare data
        data = [['Name', 'Email', 'Request', 'Public', 'Prayed', 'Date']]
        
        for prayer in queryset:
            data.append([
                prayer.name or 'Anonymous',
                prayer.email or '',
                prayer.request[:100] + '...' if len(prayer.request) > 100 else prayer.request,
                'Yes' if prayer.is_public else 'No',
                'Yes' if prayer.is_prayed_for else 'No',
                prayer.created_at.strftime('%Y-%m-%d')
            ])
        
        # Create table
        table = Table(data, colWidths=[1.2*inch, 1.8*inch, 3*inch, 0.6*inch, 0.6*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        doc.build(story)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="prayer_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = PrayerRequest.objects.all()
        status = self.request.GET.get('status')
        if status == 'open':
            queryset = queryset.filter(is_prayed_for=False)
        elif status == 'prayed':
            queryset = queryset.filter(is_prayed_for=True)
        return queryset.order_by('-created_at')


class PrayerRequestExportCardsView(StaffRequiredMixin, View):
    """Export prayer requests as beautifully designed prayer cards (PDF)."""
    def get(self, request, *args, **kwargs):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
        except ImportError:
            return HttpResponseBadRequest('PDF export requires reportlab. Please install it: pip install reportlab')
        
        queryset = self._get_queryset()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        
        # Custom styles for prayer cards
        title_style = ParagraphStyle(
            'CardTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        request_style = ParagraphStyle(
            'RequestText',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=15,
            alignment=TA_JUSTIFY,
            leading=18,
            leftIndent=20,
            rightIndent=20
        )
        
        name_style = ParagraphStyle(
            'NameText',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#4b5563'),
            alignment=TA_RIGHT,
            spaceBefore=10
        )
        
        date_style = ParagraphStyle(
            'DateText',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER,
            spaceBefore=5
        )
        
        story = []
        
        # Check if queryset is empty
        if not queryset.exists():
            return HttpResponseBadRequest('No prayer requests found to export.')
        
        for idx, prayer in enumerate(queryset):
            if idx > 0:
                story.append(PageBreak())
            
            # Decorative border/header
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph("PRAYER REQUEST", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Prayer request text - escape HTML and convert newlines
            import html
            request_text = html.escape(prayer.request).replace('\n', '<br/>')
            story.append(Paragraph(f'<b>"{request_text}"</b>', request_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Name and status - escape HTML
            name = html.escape(prayer.name or 'Anonymous')
            name_text = f"— {name}"
            if prayer.is_prayed_for:
                name_text += " ✓ (Prayed For)"
            story.append(Paragraph(name_text, name_style))
            
            # Date
            date_text = prayer.created_at.strftime('%B %d, %Y')
            story.append(Paragraph(date_text, date_style))
            
            # Footer note
            story.append(Spacer(1, 0.2*inch))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#9ca3af'),
                alignment=TA_CENTER
            )
            story.append(Paragraph("Uplift Your Morning - Prayer Request", footer_style))
        
        doc.build(story)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="prayer_cards_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = PrayerRequest.objects.all()
        status = self.request.GET.get('status')
        if status == 'open':
            queryset = queryset.filter(is_prayed_for=False)
        elif status == 'prayed':
            queryset = queryset.filter(is_prayed_for=True)
        return queryset.order_by('-created_at')


# ==================== DONATIONS ====================

class DonationListView(StaffRequiredMixin, ListView):
    """List all donations."""
    model = Donation
    template_name = 'admin/donations/list.html'
    context_object_name = 'donations'
    paginate_by = 20

    def get_queryset(self):
        queryset = Donation.objects.all()
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(email__icontains=search) |
                Q(paystack_reference__icontains=search)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = Donation.objects.count()
        context['successful'] = Donation.objects.filter(status=Donation.STATUS_SUCCESS)
        context['successful_count'] = context['successful'].count()
        context['successful_total'] = context['successful'].aggregate(
            total=Sum('amount_ghs')
        )['total'] or 0
        context['pending_count'] = Donation.objects.filter(status=Donation.STATUS_PENDING).count()
        context['failed_count'] = Donation.objects.filter(status=Donation.STATUS_FAILED).count()
        return context


class DonationVerifyView(StaffRequiredMixin, View):
    """Manually verify a donation with Paystack."""
    def post(self, request, *args, **kwargs):
        try:
            donation = Donation.objects.get(pk=kwargs['pk'])
        except Donation.DoesNotExist:
            messages.error(request, 'Donation not found.')
            return redirect('manage:donations_list')

        paystack_secret_key = getattr(settings, 'PAYSTACK_SECRET_KEY', '')
        if not paystack_secret_key:
            messages.error(request, 'Payment configuration is missing.')
            return redirect('manage:donations_list')

        headers = {
            'Authorization': f'Bearer {paystack_secret_key}',
        }

        try:
            verify_resp = requests.get(
                f'https://api.paystack.co/transaction/verify/{donation.paystack_reference}',
                headers=headers,
                timeout=30
            )
            verify_data = verify_resp.json()
        except Exception as e:
            messages.error(request, f'Could not verify payment: {str(e)}')
            return redirect('manage:donations_list')

        # Check if the API call was successful
        if verify_resp.status_code != 200:
            messages.error(request, f'Paystack API error: {verify_data.get("message", "Unknown error")}')
            return redirect('manage:donations_list')

        # Check transaction status
        transaction_data = verify_data.get('data', {})
        transaction_status = transaction_data.get('status')
        
        donation.raw_response = transaction_data
        
        if transaction_status == 'success':
            donation.status = Donation.STATUS_SUCCESS
            messages.success(request, f'Donation verified successfully! Status updated to Successful. Amount: GHS {donation.amount_ghs}')
        elif transaction_status in ['failed', 'reversed', 'insufficient_funds']:
            donation.status = Donation.STATUS_FAILED
            messages.warning(request, f'Payment verification shows status: {transaction_status}. Status updated to Failed.')
        else:
            # Still pending or other status
            messages.info(request, f'Payment status is still: {transaction_status}. Donation remains as Pending.')
        
        donation.save(update_fields=['status', 'raw_response', 'updated_at'])

        return redirect('manage:donations_list')


class DonationDetailView(StaffRequiredMixin, DetailView):
    """View donation details."""
    model = Donation
    template_name = 'admin/donations/detail.html'
    context_object_name = 'donation'


# ==================== PLEDGES ====================

class PledgeListView(StaffRequiredMixin, ListView):
    """List all pledges."""
    model = Pledge
    template_name = 'admin/pledges/list.html'
    context_object_name = 'pledges'
    paginate_by = 20

    def get_queryset(self):
        queryset = Pledge.objects.all()
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(country__icontains=search)
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = Pledge.objects.all()
        
        # Basic counts
        context['total'] = queryset.count()
        context['pending_count'] = queryset.filter(status=Pledge.STATUS_PENDING).count()
        context['confirmed_count'] = queryset.filter(status=Pledge.STATUS_CONFIRMED).count()
        context['completed_count'] = queryset.filter(status=Pledge.STATUS_COMPLETED).count()
        context['cancelled_count'] = queryset.filter(status=Pledge.STATUS_CANCELLED).count()
        
        # Analytics: Total pledge amounts by currency (only monetary pledges)
        total_by_currency = {}
        monetary_pledges = queryset.filter(pledge_type=Pledge.PLEDGE_TYPE_MONETARY)
        for pledge in monetary_pledges:
            if pledge.amount:
                currency = pledge.get_currency_display_value()
                if currency not in total_by_currency:
                    total_by_currency[currency] = 0
                total_by_currency[currency] += float(pledge.amount)
        context['total_by_currency'] = total_by_currency
        
        # Count non-monetary pledges
        context['monetary_count'] = monetary_pledges.count()
        context['non_monetary_count'] = queryset.filter(pledge_type=Pledge.PLEDGE_TYPE_NON_MONETARY).count()
        
        # Analytics: Pledges by country (convert country codes to names)
        country_stats_raw = queryset.exclude(country='').values('country').annotate(
            count=Count('id'),
            total_amount=Sum('amount')
        ).order_by('-count')[:10]  # Top 10 countries
        
        # Convert country codes to names
        from django_countries import countries as countries_dict
        country_stats = []
        for stat in country_stats_raw:
            country_name = dict(countries_dict).get(stat['country'], stat['country'])
            country_stats.append({
                'country': country_name,
                'count': stat['count'],
                'total_amount': stat['total_amount'] or 0
            })
        context['country_stats'] = country_stats
        
        # Total pledge amount (all currencies combined - approximate)
        context['total_pledge_amount'] = sum(total_by_currency.values())
        
        # Total USD amount (sum of all usd_amount fields)
        total_usd = monetary_pledges.filter(usd_amount__isnull=False).aggregate(
            total=Sum('usd_amount')
        )['total'] or 0
        context['total_usd_amount'] = total_usd
        
        context['current_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class PledgeDetailView(StaffRequiredMixin, DetailView):
    """View pledge details."""
    model = Pledge
    template_name = 'admin/pledges/detail.html'
    context_object_name = 'pledge'


class PledgeUpdateStatusView(StaffRequiredMixin, View):
    """Update the status of a pledge."""
    def post(self, request, *args, **kwargs):
        try:
            pledge = Pledge.objects.get(pk=kwargs['pk'])
            new_status = request.POST.get('status')
            admin_notes = request.POST.get('admin_notes', '')
            completed_date = request.POST.get('completed_date', '')
            
            if new_status in dict(Pledge.STATUS_CHOICES):
                pledge.status = new_status
                if admin_notes:
                    pledge.admin_notes = admin_notes
                if completed_date and new_status == Pledge.STATUS_COMPLETED:
                    from datetime import datetime
                    try:
                        pledge.completed_date = datetime.strptime(completed_date, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                pledge.save()
                messages.success(request, f'Pledge status updated to {pledge.get_status_display()}!')
            else:
                messages.error(request, 'Invalid status selected.')
        except Pledge.DoesNotExist:
            messages.error(request, 'Pledge not found.')
        return redirect('manage:pledges_detail', pk=kwargs['pk'])


class PledgeUpdateUSDView(StaffRequiredMixin, View):
    """Update the USD amount manually or trigger automatic conversion."""
    def post(self, request, *args, **kwargs):
        try:
            pledge = Pledge.objects.get(pk=kwargs['pk'])
            usd_amount = request.POST.get('usd_amount', '').strip()
            convert_auto = request.POST.get('convert_auto') == 'true'
            
            if convert_auto:
                # Trigger automatic conversion
                result = pledge.convert_to_usd()
                if result is not None:
                    pledge.save(update_fields=['usd_amount', 'conversion_rate', 'conversion_date', 'conversion_source'])
                    source_info = f" (Rate: {pledge.conversion_rate}, Source: {pledge.conversion_source})" if pledge.conversion_source else ""
                    messages.success(request, f'USD amount automatically converted: ${pledge.usd_amount:,.2f}{source_info}')
                else:
                    messages.warning(request, 'Automatic conversion failed. Please enter USD amount manually.')
            elif usd_amount:
                # Manual entry
                try:
                    from decimal import Decimal
                    usd_value = Decimal(usd_amount)
                    if usd_value < 0:
                        messages.error(request, 'USD amount cannot be negative.')
                    else:
                        pledge.usd_amount = usd_value
                        pledge.save(update_fields=['usd_amount'])
                        messages.success(request, f'USD amount updated to ${pledge.usd_amount:,.2f}')
                except (ValueError, Exception) as e:
                    messages.error(request, f'Invalid USD amount: {str(e)}')
            else:
                # Clear USD amount
                pledge.usd_amount = None
                pledge.save(update_fields=['usd_amount'])
                messages.success(request, 'USD amount cleared.')
                
        except Pledge.DoesNotExist:
            messages.error(request, 'Pledge not found.')
        except Exception as e:
            messages.error(request, f'Error updating USD amount: {str(e)}')
        
        return redirect('manage:pledges_detail', pk=kwargs['pk'])


class PledgeDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a pledge."""
    model = Pledge
    template_name = 'admin/pledges/delete.html'
    success_url = reverse_lazy('manage:pledges_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Pledge deleted successfully!')
        return super().delete(request, *args, **kwargs)


class PledgeExportCSVView(StaffRequiredMixin, View):
    """Export pledges as CSV."""
    def get(self, request, *args, **kwargs):
        queryset = self._get_queryset()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="pledges_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Email', 'Phone', 'Country', 'Preferred Contact Method', 'Contact Info', 'Pledge Type', 'Amount', 'Currency', 'Donation Frequency', 'Custom Frequency', 'Redemption Date', 'Redemption Timeframe', 'Non-Monetary Description', 'Status', 'Date Submitted', 'Completed Date', 'Additional Notes'])
        
        for pledge in queryset:
            frequency_display = pledge.get_donation_frequency_display() if pledge.donation_frequency else ''
            if pledge.donation_frequency == Pledge.FREQUENCY_CUSTOM:
                frequency_display = f"{frequency_display} - {pledge.custom_frequency}"
            writer.writerow([
                pledge.full_name,
                pledge.email,
                pledge.phone or '',
                pledge.get_country_name() or '',
                pledge.get_preferred_contact_method_display(),
                pledge.contact_info or '',
                pledge.get_pledge_type_display(),
                pledge.amount if pledge.pledge_type == Pledge.PLEDGE_TYPE_MONETARY else '',
                pledge.get_currency_display_value() if pledge.pledge_type == Pledge.PLEDGE_TYPE_MONETARY else '',
                frequency_display,
                pledge.custom_frequency if pledge.donation_frequency == Pledge.FREQUENCY_CUSTOM else '',
                pledge.redemption_date.strftime('%Y-%m-%d') if pledge.redemption_date else '',
                pledge.redemption_timeframe or '',
                pledge.non_monetary_description if pledge.pledge_type == Pledge.PLEDGE_TYPE_NON_MONETARY else '',
                pledge.get_status_display(),
                pledge.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                pledge.completed_date.strftime('%Y-%m-%d') if pledge.completed_date else '',
                pledge.additional_notes or ''
            ])
        
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = Pledge.objects.all()
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(country__icontains=search)
            )
        return queryset.order_by('-created_at')


class PledgeExportExcelView(StaffRequiredMixin, View):
    """Export pledges as Excel."""
    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponseBadRequest('Excel export requires openpyxl. Please install it: pip install openpyxl')
        
        queryset = self._get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Pledges"
        
        # Header row
        headers = ['Name', 'Email', 'Phone', 'Country', 'Preferred Contact Method', 'Contact Info', 'Pledge Type', 'Amount', 'Currency', 'Donation Frequency', 'Custom Frequency', 'Redemption Date', 'Redemption Timeframe', 'Non-Monetary Description', 'Status', 'Date Submitted', 'Completed Date', 'Additional Notes']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, pledge in enumerate(queryset, 2):
            frequency_display = pledge.get_donation_frequency_display() if pledge.donation_frequency else ''
            if pledge.donation_frequency == Pledge.FREQUENCY_CUSTOM:
                frequency_display = f"{frequency_display} - {pledge.custom_frequency}"
            ws.cell(row=row_num, column=1, value=pledge.full_name)
            ws.cell(row=row_num, column=2, value=pledge.email)
            ws.cell(row=row_num, column=3, value=pledge.phone or '')
            ws.cell(row=row_num, column=4, value=pledge.get_country_name() or '')
            ws.cell(row=row_num, column=5, value=pledge.get_preferred_contact_method_display())
            ws.cell(row=row_num, column=6, value=pledge.contact_info or '')
            ws.cell(row=row_num, column=7, value=pledge.get_pledge_type_display())
            ws.cell(row=row_num, column=8, value=float(pledge.amount) if pledge.pledge_type == Pledge.PLEDGE_TYPE_MONETARY and pledge.amount else '')
            ws.cell(row=row_num, column=9, value=pledge.get_currency_display_value() if pledge.pledge_type == Pledge.PLEDGE_TYPE_MONETARY else '')
            ws.cell(row=row_num, column=10, value=frequency_display)
            ws.cell(row=row_num, column=11, value=pledge.custom_frequency if pledge.donation_frequency == Pledge.FREQUENCY_CUSTOM else '')
            ws.cell(row=row_num, column=12, value=pledge.redemption_date.strftime('%Y-%m-%d') if pledge.redemption_date else '')
            ws.cell(row=row_num, column=13, value=pledge.redemption_timeframe or '')
            ws.cell(row=row_num, column=14, value=pledge.non_monetary_description if pledge.pledge_type == Pledge.PLEDGE_TYPE_NON_MONETARY else '')
            ws.cell(row=row_num, column=15, value=pledge.get_status_display())
            ws.cell(row=row_num, column=16, value=pledge.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=17, value=pledge.completed_date.strftime('%Y-%m-%d') if pledge.completed_date else '')
            ws.cell(row=row_num, column=18, value=pledge.additional_notes or '')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 20
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 50
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="pledges_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = Pledge.objects.all()
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(country__icontains=search)
            )
        return queryset.order_by('-created_at')


class PledgeFindDuplicatesView(StaffRequiredMixin, View):
    """Find duplicate pledges based on email or name+email combination."""
    def get(self, request, *args, **kwargs):
        from django.db.models import Count
        from collections import defaultdict
        
        # Find duplicates by email
        email_duplicates = Pledge.objects.values('email').annotate(
            count=Count('id')
        ).filter(count__gt=1)
        
        # Find duplicates by name + email combination
        name_email_duplicates = Pledge.objects.values('full_name', 'email').annotate(
            count=Count('id')
        ).filter(count__gt=1)
        
        # Group duplicates
        duplicate_groups = defaultdict(list)
        
        for dup in email_duplicates:
            pledges = Pledge.objects.filter(email=dup['email']).order_by('created_at')
            if pledges.count() > 1:
                duplicate_groups[dup['email']] = list(pledges)
        
        # Also check name+email combinations
        for dup in name_email_duplicates:
            key = f"{dup['full_name']}|{dup['email']}"
            if key not in duplicate_groups:
                pledges = Pledge.objects.filter(
                    full_name=dup['full_name'],
                    email=dup['email']
                ).order_by('created_at')
                if pledges.count() > 1:
                    duplicate_groups[key] = list(pledges)
        
        # Convert to list of dicts for better template handling
        duplicate_groups_list = []
        for key, pledges_list in duplicate_groups.items():
            duplicate_groups_list.append({
                'key': key,
                'pledges': pledges_list,
                'name': key.split('|')[0] if '|' in key else key,
                'email': key.split('|')[1] if '|' in key else key,
            })
        
        context = {
            'duplicate_groups': duplicate_groups_list,
            'total_duplicates': sum(len(group['pledges']) - 1 for group in duplicate_groups_list),
        }
        
        return render(request, 'admin/pledges/find_duplicates.html', context)
    
    def post(self, request, *args, **kwargs):
        """Delete selected duplicate pledges."""
        selected_ids = request.POST.getlist('pledge_ids')
        removed_count = 0
        
        if selected_ids:
            for pledge_id in selected_ids:
                try:
                    pledge = Pledge.objects.get(pk=pledge_id)
                    pledge.delete()
                    removed_count += 1
                except Pledge.DoesNotExist:
                    pass
        
        messages.success(request, f'Successfully deleted {removed_count} duplicate pledge(s).')
        return redirect('manage:pledges_find_duplicates')


class PledgeRemoveDuplicatesView(StaffRequiredMixin, View):
    """Remove duplicate pledges, keeping the oldest one."""
    def post(self, request, *args, **kwargs):
        from django.db.models import Count
        from django.contrib import messages
        
        removed_count = 0
        
        # Find duplicates by email
        email_duplicates = Pledge.objects.values('email').annotate(
            count=Count('id')
        ).filter(count__gt=1)
        
        for dup in email_duplicates:
            pledges = list(Pledge.objects.filter(email=dup['email']).order_by('created_at'))
            if len(pledges) > 1:
                # Keep the first (oldest) one, delete the rest
                for pledge in pledges[1:]:
                    pledge.delete()
                    removed_count += 1
        
        # Also check name+email combinations
        name_email_duplicates = Pledge.objects.values('full_name', 'email').annotate(
            count=Count('id')
        ).filter(count__gt=1)
        
        for dup in name_email_duplicates:
            pledges = list(Pledge.objects.filter(
                full_name=dup['full_name'],
                email=dup['email']
            ).order_by('created_at'))
            if len(pledges) > 1:
                # Keep the first (oldest) one, delete the rest
                for pledge in pledges[1:]:
                    pledge.delete()
                    removed_count += 1
        
        messages.success(request, f'Successfully removed {removed_count} duplicate pledge(s).')
        return redirect('manage:pledges_list')


class PledgeConvertToUSDView(StaffRequiredMixin, View):
    """Convert all existing pledges to USD (update usd_amount field)."""
    def post(self, request, *args, **kwargs):
        import logging
        logger = logging.getLogger(__name__)
        
        updated_count = 0
        failed_count = 0
        failed_details = []
        
        monetary_pledges = Pledge.objects.filter(
            pledge_type=Pledge.PLEDGE_TYPE_MONETARY,
            amount__isnull=False
        )
        
        for pledge in monetary_pledges:
            try:
                # Get currency code for logging
                if pledge.currency == Pledge.CURRENCY_OTHER:
                    currency_code = pledge.other_currency or 'UNKNOWN'
                else:
                    currency_code = pledge.currency or 'UNKNOWN'
                
                logger.info(f"Processing pledge {pledge.id}: amount={pledge.amount}, currency={currency_code}, pledge_type={pledge.pledge_type}")
                
                # Convert to USD
                result = pledge.convert_to_usd()
                
                if result is not None:
                    pledge.save(update_fields=['usd_amount', 'conversion_rate', 'conversion_date', 'conversion_source'])
                    updated_count += 1
                    logger.info(f"✓ Converted pledge {pledge.id}: {pledge.amount} {currency_code} = ${result} USD (rate: {pledge.conversion_rate}, source: {pledge.conversion_source})")
                else:
                    failed_count += 1
                    error_detail = f"Pledge {pledge.id} ({currency_code}): Conversion returned None"
                    failed_details.append(error_detail)
                    logger.warning(f"✗ Conversion failed for pledge {pledge.id}: {error_detail}")
            except Exception as e:
                failed_count += 1
                error_msg = f"Pledge {pledge.id}: {str(e)}"
                failed_details.append(error_msg)
                logger.error(f"✗ Exception converting pledge {pledge.id}: {e}", exc_info=True)
        
        if failed_count > 0:
            # Show first few error details in the message
            error_preview = failed_details[:3]  # Show first 3 errors
            error_msg = f'Converted {updated_count} pledge(s) to USD. {failed_count} failed.'
            if error_preview:
                error_msg += f' Issues: {"; ".join(error_preview)}'
                if len(failed_details) > 3:
                    error_msg += f' (and {len(failed_details) - 3} more)'
            messages.warning(request, error_msg)
        else:
            messages.success(
                request,
                f'Successfully converted {updated_count} pledge(s) to USD.'
            )
        
        return redirect('manage:pledges_list')


# ==================== TESTIMONIES ====================

class TestimonyListView(StaffRequiredMixin, ListView):
    """List all testimonies."""
    model = Testimony
    template_name = 'admin/testimonies/list.html'
    context_object_name = 'testimonies'
    paginate_by = 20

    def get_queryset(self):
        queryset = Testimony.objects.all()
        # Filter by approval status
        status = self.request.GET.get('status')
        if status == 'approved':
            queryset = queryset.filter(is_approved=True)
        elif status == 'pending':
            queryset = queryset.filter(is_approved=False)
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = Testimony.objects.count()
        context['approved_count'] = Testimony.objects.filter(is_approved=True).count()
        context['pending_count'] = Testimony.objects.filter(is_approved=False).count()
        return context


class TestimonyDetailView(StaffRequiredMixin, DetailView):
    """View details of a testimony."""
    model = Testimony
    template_name = 'admin/testimonies/detail.html'
    context_object_name = 'testimony'


class TestimonyApproveView(StaffRequiredMixin, View):
    """Approve a testimony."""
    def post(self, request, *args, **kwargs):
        try:
            testimony = Testimony.objects.get(pk=kwargs['pk'])
            testimony.is_approved = True
            testimony.save()
            messages.success(request, 'Testimony approved!')
            return redirect('manage:testimonies_detail', pk=testimony.pk)
        except Testimony.DoesNotExist:
            messages.error(request, 'Testimony not found.')
            return redirect('manage:testimonies_list')


class TestimonyDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a testimony."""
    model = Testimony
    template_name = 'admin/testimonies/delete.html'
    success_url = reverse_lazy('manage:testimonies_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Testimony deleted successfully!')
        return super().delete(request, *args, **kwargs)


class TestimonyExportCSVView(StaffRequiredMixin, View):
    """Export testimonies as CSV."""
    def get(self, request, *args, **kwargs):
        queryset = self._get_queryset()
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="testimonies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Country', 'Testimony', 'Is Approved', 'Is Featured', 'Date Submitted'])
        
        for testimony in queryset:
            writer.writerow([
                testimony.name or 'Anonymous',
                testimony.country or '',
                testimony.testimony,
                'Yes' if testimony.is_approved else 'No',
                'Yes' if testimony.featured else 'No',
                testimony.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = Testimony.objects.all()
        status = self.request.GET.get('status')
        if status == 'approved':
            queryset = queryset.filter(is_approved=True)
        elif status == 'pending':
            queryset = queryset.filter(is_approved=False)
        return queryset.order_by('-created_at')


class TestimonyExportExcelView(StaffRequiredMixin, View):
    """Export testimonies as Excel."""
    def get(self, request, *args, **kwargs):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponseBadRequest('Excel export requires openpyxl. Please install it: pip install openpyxl')
        
        queryset = self._get_queryset()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Testimonies"
        
        # Header row
        headers = ['Name', 'Country', 'Testimony', 'Is Approved', 'Is Featured', 'Date Submitted']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, testimony in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=testimony.name or 'Anonymous')
            ws.cell(row=row_num, column=2, value=testimony.country or '')
            ws.cell(row=row_num, column=3, value=testimony.testimony)
            ws.cell(row=row_num, column=4, value='Yes' if testimony.is_approved else 'No')
            ws.cell(row=row_num, column=5, value='Yes' if testimony.featured else 'No')
            ws.cell(row=row_num, column=6, value=testimony.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="testimonies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = Testimony.objects.all()
        status = self.request.GET.get('status')
        if status == 'approved':
            queryset = queryset.filter(is_approved=True)
        elif status == 'pending':
            queryset = queryset.filter(is_approved=False)
        return queryset.order_by('-created_at')


class TestimonyExportPDFView(StaffRequiredMixin, View):
    """Export testimonies as PDF (spreadsheet format)."""
    def get(self, request, *args, **kwargs):
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            return HttpResponseBadRequest('PDF export requires reportlab. Please install it: pip install reportlab')
        
        queryset = self._get_queryset()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        story.append(Paragraph("Testimonies Export", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Prepare data
        data = [['Name', 'Country', 'Testimony', 'Approved', 'Featured', 'Date']]
        
        for testimony in queryset:
            data.append([
                testimony.name or 'Anonymous',
                testimony.country or '',
                testimony.testimony[:100] + '...' if len(testimony.testimony) > 100 else testimony.testimony,
                'Yes' if testimony.is_approved else 'No',
                'Yes' if testimony.featured else 'No',
                testimony.created_at.strftime('%Y-%m-%d')
            ])
        
        # Create table
        table = Table(data, colWidths=[1.2*inch, 1*inch, 3.5*inch, 0.8*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        
        story.append(table)
        doc.build(story)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="testimonies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _get_queryset(self):
        """Get filtered queryset based on request parameters."""
        queryset = Testimony.objects.all()
        status = self.request.GET.get('status')
        if status == 'approved':
            queryset = queryset.filter(is_approved=True)
        elif status == 'pending':
            queryset = queryset.filter(is_approved=False)
        return queryset.order_by('-created_at')


# ==================== 40 DAYS CONFIGURATION ====================

class FortyDaysConfigListView(StaffRequiredMixin, ListView):
    """List all 40 Days configurations."""
    model = FortyDaysConfig
    template_name = 'admin/fortydays/list.html'
    context_object_name = 'configs'
    paginate_by = 20

    def get_queryset(self):
        queryset = FortyDaysConfig.objects.all()
        # Filter by active status
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        return queryset.order_by('-start_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = FortyDaysConfig.objects.count()
        context['active_count'] = FortyDaysConfig.objects.filter(is_active=True).count()
        context['inactive_count'] = FortyDaysConfig.objects.filter(is_active=False).count()
        return context


class FortyDaysConfigCreateView(StaffRequiredMixin, CreateView):
    """Create a new 40 Days configuration."""
    model = FortyDaysConfig
    template_name = 'admin/fortydays/form.html'
    fields = [
        'start_date', 'end_date', 'is_active', 'banner_image',
        'morning_youtube_url', 'morning_facebook_url',
        'evening_youtube_url', 'evening_facebook_url'
    ]
    success_url = reverse_lazy('manage:fortydays_list')

    def form_valid(self, form):
        messages.success(self.request, '40 Days configuration created successfully!')
        return super().form_valid(form)


class FortyDaysConfigUpdateView(StaffRequiredMixin, UpdateView):
    """Update a 40 Days configuration."""
    model = FortyDaysConfig
    template_name = 'admin/fortydays/form.html'
    fields = [
        'start_date', 'end_date', 'is_active', 'banner_image',
        'morning_youtube_url', 'morning_facebook_url',
        'evening_youtube_url', 'evening_facebook_url'
    ]
    success_url = reverse_lazy('manage:fortydays_list')

    def form_valid(self, form):
        # Save the form (this handles image upload if present)
        response = super().form_valid(form)
        messages.success(self.request, '40 Days configuration updated successfully!')
        return response


class FortyDaysConfigDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a 40 Days configuration."""
    model = FortyDaysConfig
    template_name = 'admin/fortydays/delete.html'
    success_url = reverse_lazy('manage:fortydays_list')

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, '40 Days configuration deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ==================== SITE SETTINGS ====================

class SiteSettingsUpdateView(StaffRequiredMixin, UpdateView):
    """Update site settings (singleton model)."""
    model = SiteSettings
    template_name = 'admin/sitesettings/form.html'
    fields = ['zoom_link']
    success_url = reverse_lazy('manage:sitesettings_edit')

    def get_object(self, queryset=None):
        # Get or create the singleton instance
        obj, created = SiteSettings.objects.get_or_create(pk=1)
        return obj

    def form_valid(self, form):
        messages.success(self.request, 'Site settings updated successfully!')
        return super().form_valid(form)


# ==================== COUNSELING BOOKINGS ====================

from .notifications import send_booking_approval_notifications, create_google_calendar_event

class CounselingBookingListView(StaffRequiredMixin, ListView):
    """List all counseling bookings."""
    model = CounselingBooking
    template_name = 'admin/counseling/list.html'
    context_object_name = 'bookings'
    paginate_by = 20

    def get_queryset(self):
        queryset = CounselingBooking.objects.all()
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                full_name__icontains=search
            ) | queryset.filter(
                email__icontains=search
            ) | queryset.filter(
                phone__icontains=search
            )
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = CounselingBooking.objects.count()
        context['pending'] = CounselingBooking.objects.filter(status=CounselingBooking.STATUS_PENDING).count()
        context['approved'] = CounselingBooking.objects.filter(status=CounselingBooking.STATUS_APPROVED).count()
        return context


class CounselingBookingDetailView(StaffRequiredMixin, DetailView):
    """View details of a counseling booking."""
    model = CounselingBooking
    template_name = 'admin/counseling/detail.html'
    context_object_name = 'booking'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        booking = context['booking']
        
        # Build previews for email and SMS
        # Use approved date/time if available, otherwise use preferred date/time for preview
        if booking.approved_date and booking.approved_time:
            context['email_preview'] = self._build_email_preview(booking)
            context['sms_preview'] = self._build_sms_preview(booking)
        elif booking.status == 'pending':
            # For pending bookings, show preview using preferred date/time
            context['email_preview'] = self._build_email_preview(booking, use_preferred=True)
            context['sms_preview'] = self._build_sms_preview(booking, use_preferred=True)
        
        return context
    
    def _build_email_preview(self, booking, use_preferred=False):
        """Build email preview text."""
        if use_preferred or not (booking.approved_date and booking.approved_time):
            date_str = booking.preferred_date.strftime('%B %d, %Y')
            time_str = booking.preferred_time.strftime('%I:%M %p')
        else:
            date_str = booking.approved_date.strftime('%B %d, %Y')
            time_str = booking.approved_time.strftime('%I:%M %p')
        
        zoom_link = 'https://us02web.zoom.us/j/6261738082?pwd=RWNTU3RsNEdGMWcxOGpxRWtNM00zdz09'
        
        return f"""Dear {booking.full_name},

Your counseling session request has been approved!

Session Details:
- Date: {date_str}
- Time: {time_str}
- Duration: {booking.duration_minutes} minutes
- Topic: {booking.topic or 'General Counseling'}

Meeting Link:
{zoom_link}

Please join the meeting at the scheduled time using the link above. If you need to reschedule or have any questions, please contact us.

We look forward to meeting with you.

Blessings,
Uplift Your Morning Team"""
    
    def _build_sms_preview(self, booking, use_preferred=False):
        """Build SMS preview text."""
        if use_preferred or not (booking.approved_date and booking.approved_time):
            date_str = booking.preferred_date.strftime('%B %d, %Y')
            time_str = booking.preferred_time.strftime('%I:%M %p')
        else:
            date_str = booking.approved_date.strftime('%B %d, %Y')
            time_str = booking.approved_time.strftime('%I:%M %p')
        
        # Keep SMS concise and short (no Zoom link - that's in email only)
        return f"Session approved! {date_str} at {time_str}. Check email for Zoom link. Uplift Your Morning"


# ==================== SUBSCRIBERS ====================

class SubscriberListView(StaffRequiredMixin, ListView):
    """List all subscribers with filters and search."""
    model = Subscriber
    template_name = 'admin/subscribers/list.html'
    context_object_name = 'subscribers'
    paginate_by = 30

    def get_queryset(self):
        queryset = Subscriber.objects.all()
        
        # Filter by channel
        channel = self.request.GET.get('channel')
        if channel:
            queryset = queryset.filter(channel=channel)
        
        # Filter by status
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        # Filter by preferences
        preference = self.request.GET.get('preference')
        if preference == 'daily':
            queryset = queryset.filter(receive_daily_devotion=True, is_active=True)
        elif preference == 'special':
            queryset = queryset.filter(receive_special_programs=True, is_active=True)
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = Subscriber.objects.count()
        context['active_count'] = Subscriber.objects.filter(is_active=True).count()
        context['inactive_count'] = Subscriber.objects.filter(is_active=False).count()
        context['email_count'] = Subscriber.objects.filter(channel=Subscriber.CHANNEL_EMAIL, is_active=True).count()
        context['whatsapp_count'] = Subscriber.objects.filter(channel=Subscriber.CHANNEL_WHATSAPP, is_active=True).count()
        context['daily_devotion_count'] = Subscriber.objects.filter(is_active=True, receive_daily_devotion=True).count()
        context['special_programs_count'] = Subscriber.objects.filter(is_active=True, receive_special_programs=True).count()
        return context


class SubscriberActivateView(StaffRequiredMixin, View):
    """Activate a subscriber."""
    def post(self, request, pk):
        subscriber = get_object_or_404(Subscriber, pk=pk)
        subscriber.is_active = True
        subscriber.save()
        messages.success(request, f'Subscriber activated: {subscriber.email or subscriber.phone}')
        return redirect('manage:subscribers_list')


class SubscriberDeactivateView(StaffRequiredMixin, View):
    """Deactivate a subscriber."""
    def post(self, request, pk):
        subscriber = get_object_or_404(Subscriber, pk=pk)
        subscriber.is_active = False
        subscriber.save()
        messages.success(request, f'Subscriber deactivated: {subscriber.email or subscriber.phone}')
        return redirect('manage:subscribers_list')


class SubscriberDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a subscriber."""
    model = Subscriber
    template_name = 'admin/subscribers/delete.html'
    success_url = reverse_lazy('manage:subscribers_list')
    
    def delete(self, request, *args, **kwargs):
        subscriber = self.get_object()
        messages.success(request, f'Subscriber deleted: {subscriber.email or subscriber.phone}')
        return super().delete(request, *args, **kwargs)


# ==================== NOTIFICATIONS ====================

class NotificationScheduleListView(StaffRequiredMixin, ListView):
    """List all scheduled notifications."""
    model = ScheduledNotification
    template_name = 'admin/notifications/list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        queryset = ScheduledNotification.objects.all()
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        return queryset.order_by('-scheduled_date', '-scheduled_time')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total'] = ScheduledNotification.objects.count()
        context['scheduled'] = ScheduledNotification.objects.filter(status=ScheduledNotification.STATUS_SCHEDULED, is_paused=False).count()
        context['paused'] = ScheduledNotification.objects.filter(is_paused=True).count()
        context['sent'] = ScheduledNotification.objects.filter(status=ScheduledNotification.STATUS_SENT).count()
        return context


class NotificationScheduleCreateView(StaffRequiredMixin, CreateView):
    """Create a new scheduled notification."""
    model = ScheduledNotification
    template_name = 'admin/notifications/form.html'
    fields = [
        'title', 'devotion', 'custom_message', 'scheduled_date', 'scheduled_time',
        'send_to_email', 'send_to_whatsapp', 'only_daily_devotion_subscribers', 'notes'
    ]
    success_url = reverse_lazy('manage:notifications_list')

    def get_initial(self):
        """Set default values for the form."""
        from datetime import date, time
        from django.utils import timezone
        
        # Get today's date
        today = date.today()
        
        # Default time: 5:00 AM
        default_time = time(5, 0)
        
        return {
            'title': 'Daily Devotion - Uplift Your Morning',
            'scheduled_date': today,
            'scheduled_time': default_time,
            'send_to_email': True,
            'send_to_sms': True,
            'send_to_whatsapp': True,
            'only_daily_devotion_subscribers': True,
        }

    def form_valid(self, form):
        messages.success(self.request, 'Notification scheduled successfully!')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get today's devotion for preview
        from datetime import date
        from apps.devotions.models import Devotion
        try:
            today_devotion = Devotion.objects.filter(
                is_published=True,
                publish_date=date.today()
            ).first()
            context['today_devotion'] = today_devotion
        except:
            context['today_devotion'] = None
        return context


class NotificationScheduleDetailView(StaffRequiredMixin, DetailView):
    """View details and preview of a scheduled notification."""
    model = ScheduledNotification
    template_name = 'admin/notifications/detail.html'
    context_object_name = 'notification'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        notification = self.get_object()
        
        # Build preview messages
        from apps.subscriptions.management.commands.send_daily_devotions import Command as DevotionCommand
        command = DevotionCommand()
        
        # Get the devotion to use
        devotion = notification.devotion
        if not devotion:
            from datetime import date
            from apps.devotions.models import Devotion
            devotion = Devotion.objects.filter(
                is_published=True,
                publish_date=notification.scheduled_date
            ).first()
        
        # Build email preview with subject
        if devotion:
            email_subject = f'{notification.title} - {devotion.title}'
            email_preview = command._build_devotion_email(devotion)
            if notification.custom_message:
                email_preview += f"\n\n{notification.custom_message}"
            sms_preview = command._build_devotion_sms(devotion)
            if notification.custom_message:
                sms_preview += f"\n\n{notification.custom_message[:100]}..."  # Truncate for SMS
            # WhatsApp preview uses short content (max 300 chars)
            whatsapp_preview = command._build_devotion_whatsapp(devotion)
            if notification.custom_message:
                remaining = 300 - len(whatsapp_preview) - 5
                if remaining > 20:
                    custom_msg = notification.custom_message[:remaining] + "..." if len(notification.custom_message) > remaining else notification.custom_message
                    whatsapp_preview += f"\n\n{custom_msg}"
                else:
                    available = 300 - len(notification.custom_message[:50]) - 10
                    whatsapp_preview = whatsapp_preview[:available] + "..."
                    whatsapp_preview += f"\n\n{notification.custom_message[:50]}"
            has_devotion = True
        else:
            email_subject = notification.title
            email_preview = command._build_no_devotion_email()
            if notification.custom_message:
                email_preview += f"\n\n{notification.custom_message}"
            sms_preview = command._build_no_devotion_sms()
            if notification.custom_message:
                sms_preview += f"\n\n{notification.custom_message[:100]}..."
            # WhatsApp preview uses short content (max 300 chars)
            whatsapp_preview = command._build_no_devotion_sms()  # Use SMS format for no devotion
            if notification.custom_message:
                remaining = 300 - len(whatsapp_preview) - 5
                if remaining > 20:
                    custom_msg = notification.custom_message[:remaining] + "..." if len(notification.custom_message) > remaining else notification.custom_message
                    whatsapp_preview += f"\n\n{custom_msg}"
            has_devotion = False
        
        # Build HTML preview with image for email and WhatsApp
        from django.conf import settings
        site_url = getattr(settings, 'SITE_URL', 'https://upliftyourmorning.com')
        email_html_preview = None
        whatsapp_html_preview = None
        
        if devotion and devotion.image:
            # Build HTML email preview with image
            # Use request to get absolute URL
            request = self.request
            if request:
                image_url = request.build_absolute_uri(devotion.image.url)
            else:
                image_url = f"{site_url}{devotion.image.url}"
            
            # Escape HTML in the title for safety, but preserve line breaks in content
            from django.utils.html import escape
            escaped_title = escape(devotion.title)
            # Convert plain text preview to HTML, preserving line breaks
            html_content = escape(email_preview).replace('\n', '<br>')
            
            email_html_preview = f"""
<div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; line-height: 1.6; color: #1f2937;">
    <div style="margin-bottom: 20px; text-align: center;">
        <img src="{image_url}" alt="{escaped_title}" style="width: 100%; max-width: 600px; height: auto; border-radius: 8px; display: block; margin: 0 auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
    </div>
    <div style="white-space: pre-wrap; padding: 10px 0;">{html_content}</div>
</div>
"""
            whatsapp_html_preview = email_html_preview  # Same for WhatsApp
        
        context['email_subject'] = email_subject
        context['email_preview'] = email_preview
        context['email_html_preview'] = email_html_preview  # HTML version with image
        context['sms_preview'] = sms_preview
        context['whatsapp_preview'] = whatsapp_preview  # Full email content for WhatsApp
        context['whatsapp_html_preview'] = whatsapp_html_preview  # HTML version with image
        context['devotion'] = devotion
        context['has_devotion'] = has_devotion
        
        # Get recipient counts
        from apps.subscriptions.models import Subscriber
        email_count = 0
        sms_count = 0
        whatsapp_count = 0
        
        if notification.send_to_email:
            email_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_EMAIL,
                is_active=True,
                email__isnull=False
            ).exclude(email='')
            if notification.only_daily_devotion_subscribers:
                email_qs = email_qs.filter(receive_daily_devotion=True)
            email_count = email_qs.count()
        
        if notification.send_to_sms:
            sms_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_SMS,
                is_active=True,
                phone__isnull=False
            ).exclude(phone='')
            if notification.only_daily_devotion_subscribers:
                sms_qs = sms_qs.filter(receive_daily_devotion=True)
            sms_count = sms_qs.count()
        
        if notification.send_to_whatsapp:
            whatsapp_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_WHATSAPP,
                is_active=True,
                phone__isnull=False
            ).exclude(phone='')
            if notification.only_daily_devotion_subscribers:
                whatsapp_qs = whatsapp_qs.filter(receive_daily_devotion=True)
            whatsapp_count = whatsapp_qs.count()
        
        context['email_recipient_count'] = email_count
        context['sms_recipient_count'] = sms_count
        context['whatsapp_recipient_count'] = whatsapp_count
        
        # Get error details from session (if notification was just sent)
        error_details = self.request.session.pop(f'notification_{notification.pk}_errors', None)
        if error_details:
            context['error_details'] = error_details
        
        # Calculate statistics for display (only if notification has been sent)
        if notification.status == 'sent':
            # Use actual attempts (sent + failed) to calculate totals, not current recipient counts
            # This ensures accuracy even if subscriber counts changed after sending
            email_attempts = notification.email_sent_count + notification.email_failed_count
            sms_attempts = notification.sms_sent_count + notification.sms_failed_count
            whatsapp_attempts = notification.whatsapp_sent_count + notification.whatsapp_failed_count
            
            total_recipients = email_attempts + sms_attempts + whatsapp_attempts
            total_sent = notification.email_sent_count + notification.sms_sent_count + notification.whatsapp_sent_count
            total_failed = notification.email_failed_count + notification.sms_failed_count + notification.whatsapp_failed_count
            success_rate = (total_sent / total_recipients * 100) if total_recipients > 0 else 0
            
            # Calculate per-channel success rates based on actual attempts (sent + failed)
            email_total = notification.email_sent_count + notification.email_failed_count
            email_rate = (notification.email_sent_count / email_total * 100) if email_total > 0 else 0
            
            sms_total = notification.sms_sent_count + notification.sms_failed_count
            sms_rate = (notification.sms_sent_count / sms_total * 100) if sms_total > 0 else 0
            
            whatsapp_total = notification.whatsapp_sent_count + notification.whatsapp_failed_count
            whatsapp_rate = (notification.whatsapp_sent_count / whatsapp_total * 100) if whatsapp_total > 0 else 0
            
            context['total_recipients'] = total_recipients
            context['total_sent'] = total_sent
            context['total_failed'] = total_failed
            context['success_rate'] = success_rate
            context['email_rate'] = email_rate
            context['sms_rate'] = sms_rate
            context['whatsapp_rate'] = whatsapp_rate
        else:
            # Set default values for unsent notifications
            context['total_recipients'] = email_count + sms_count + whatsapp_count
            context['total_sent'] = 0
            context['total_failed'] = 0
            context['success_rate'] = 0
            context['email_rate'] = 0
            context['sms_rate'] = 0
            context['whatsapp_rate'] = 0
        
        return context


class NotificationPauseView(StaffRequiredMixin, View):
    """Pause a scheduled notification."""
    def post(self, request, pk):
        notification = get_object_or_404(ScheduledNotification, pk=pk)
        notification.pause()
        messages.success(request, f'Notification "{notification.title}" has been paused.')
        return redirect('manage:notifications_detail', pk=pk)


class NotificationResumeView(StaffRequiredMixin, View):
    """Resume a paused notification."""
    def post(self, request, pk):
        notification = get_object_or_404(ScheduledNotification, pk=pk)
        notification.resume()
        messages.success(request, f'Notification "{notification.title}" has been resumed.')
        return redirect('manage:notifications_detail', pk=pk)


class NotificationSendNowView(StaffRequiredMixin, View):
    """Manually send a notification immediately."""
    def post(self, request, pk):
        notification = get_object_or_404(ScheduledNotification, pk=pk)
        
        # Check if already sent
        if notification.status == ScheduledNotification.STATUS_SENT:
            messages.warning(request, f'Notification "{notification.title}" has already been sent.')
            return redirect('manage:notifications_detail', pk=pk)
        
        # Use the same logic as the command
        from apps.subscriptions.management.commands.send_daily_devotions import Command as DevotionCommand
        command = DevotionCommand()
        
        # Get the devotion to use (use today's date for manual sends, or scheduled date)
        from datetime import date
        from apps.devotions.models import Devotion
        
        devotion = notification.devotion
        if not devotion:
            # For manual sends, try today's devotion first, then scheduled date
            devotion = Devotion.objects.filter(
                is_published=True,
                publish_date=date.today()
            ).first()
            
            if not devotion:
                # Fall back to scheduled date
                devotion = Devotion.objects.filter(
                    is_published=True,
                    publish_date=notification.scheduled_date
                ).first()
        
        # Check if devotion exists
        if not devotion:
            messages.error(
                request, 
                f'No published devotion found for today or scheduled date ({notification.scheduled_date}). '
                'Please publish a devotion first or link a specific devotion to this notification.'
            )
            return redirect('manage:notifications_detail', pk=pk)
        
        # Build messages
        email_subject = f'{notification.title} - {devotion.title}' if devotion else notification.title
        email_message = command._build_devotion_email(devotion)
        if notification.custom_message:
            email_message += f"\n\n{notification.custom_message}"
        
        sms_message = command._build_devotion_sms(devotion)
        if notification.custom_message:
            sms_message += f"\n\n{notification.custom_message[:100]}..."
        
        # WhatsApp gets short content (max 300 chars)
        whatsapp_message = command._build_devotion_whatsapp(devotion)
        if notification.custom_message:
            # Add custom message but ensure total stays under 300 chars
            remaining = 300 - len(whatsapp_message) - 5
            if remaining > 20:
                custom_msg = notification.custom_message[:remaining] + "..." if len(notification.custom_message) > remaining else notification.custom_message
                whatsapp_message += f"\n\n{custom_msg}"
            else:
                # Truncate whatsapp message to make room
                available = 300 - len(notification.custom_message[:50]) - 10
                whatsapp_message = whatsapp_message[:available] + "..."
                whatsapp_message += f"\n\n{notification.custom_message[:50]}"
        
        # Get recipients
        from apps.subscriptions.models import Subscriber
        email_subscribers = []
        sms_subscribers = []
        whatsapp_subscribers = []
        
        if notification.send_to_email:
            email_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_EMAIL,
                is_active=True,
                email__isnull=False
            ).exclude(email='')
            if notification.only_daily_devotion_subscribers:
                email_qs = email_qs.filter(receive_daily_devotion=True)
            email_subscribers = list(email_qs)
        
        if notification.send_to_sms:
            sms_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_SMS,
                is_active=True,
                phone__isnull=False
            ).exclude(phone='')
            if notification.only_daily_devotion_subscribers:
                sms_qs = sms_qs.filter(receive_daily_devotion=True)
            sms_subscribers = list(sms_qs)
        
        if notification.send_to_whatsapp:
            whatsapp_qs = Subscriber.objects.filter(
                channel=Subscriber.CHANNEL_WHATSAPP,
                is_active=True,
                phone__isnull=False
            ).exclude(phone='')
            if notification.only_daily_devotion_subscribers:
                whatsapp_qs = whatsapp_qs.filter(receive_daily_devotion=True)
            whatsapp_subscribers = list(whatsapp_qs)
        
        if not email_subscribers and not sms_subscribers and not whatsapp_subscribers:
            messages.warning(request, 'No active subscribers found for the selected channels and filters.')
            return redirect('manage:notifications_detail', pk=pk)
        
        # Send emails
        email_sent = 0
        email_failed = 0
        email_errors = {}
        if email_subscribers:
            from django.core.mail import send_mail
            
            # Check email configuration first
            if settings.EMAIL_BACKEND == 'django.core.mail.backends.console.EmailBackend':
                messages.warning(
                    request,
                    '⚠️ Email backend is set to console. Emails will only print to console, not actually send. '
                    'Please configure EMAIL_BACKEND in .env file.'
                )
            
            if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
                messages.warning(
                    request,
                    '⚠️ Email credentials not configured. Please set EMAIL_HOST_USER and EMAIL_HOST_PASSWORD in .env file.'
                )
            
            for subscriber in email_subscribers:
                try:
                    send_mail(
                        email_subject,
                        email_message,
                        settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@upliftyourmorning.com',
                        [subscriber.email],
                        fail_silently=False,
                    )
                    email_sent += 1
                except Exception as e:
                    email_failed += 1
                    error_msg = str(e)
                    # Group errors by type
                    if error_msg not in email_errors:
                        email_errors[error_msg] = []
                    email_errors[error_msg].append(subscriber.email)
        
        # Display grouped email errors
        if email_errors:
            for error_msg, emails in email_errors.items():
                if len(emails) > 3:
                    messages.error(
                        request,
                        f'❌ Email Error ({len(emails)} recipients): {error_msg}'
                    )
                else:
                    for email in emails[:3]:  # Limit to 3 to avoid spam
                        messages.error(request, f'❌ {email}: {error_msg}')
        
        # Send SMS (via FastR API - short messages)
        sms_sent = 0
        sms_failed = 0
        sms_errors = {}
        if sms_subscribers:
            for subscriber in sms_subscribers:
                try:
                    result = command._send_sms(subscriber.phone, sms_message)
                    # Only count as sent if _send_sms returned True (actually sent)
                    # If False, SMS is not configured and we skip silently
                    if result is True:
                        sms_sent += 1
                    # If result is False, SMS not configured - skip silently (don't count as sent or failed)
                except Exception as e:
                    sms_failed += 1
                    error_msg = str(e)
                    if error_msg not in sms_errors:
                        sms_errors[error_msg] = []
                    sms_errors[error_msg].append(subscriber.phone)
        
        # Send WhatsApp (via Twilio API - full email content)
        whatsapp_sent = 0
        whatsapp_failed = 0
        whatsapp_errors = {}
        if whatsapp_subscribers:
            from apps.subscriptions.whatsapp import send_whatsapp_message
            for subscriber in whatsapp_subscribers:
                try:
                    # WhatsApp gets the full devotion email content
                    send_whatsapp_message(subscriber.phone, whatsapp_message)
                    whatsapp_sent += 1
                except Exception as e:
                    whatsapp_failed += 1
                    error_msg = str(e)
                    if error_msg not in whatsapp_errors:
                        whatsapp_errors[error_msg] = []
                    whatsapp_errors[error_msg].append(subscriber.phone)
        
        # Display SMS errors
        if sms_errors:
            for error_msg, phones in sms_errors.items():
                if len(phones) > 3:
                    messages.error(
                        request,
                        f'❌ SMS Error ({len(phones)} recipients): {error_msg}'
                    )
                else:
                    for phone in phones[:3]:  # Limit to 3 to avoid spam
                        messages.error(request, f'❌ SMS {phone}: {error_msg}')
        
        # Display WhatsApp errors
        if whatsapp_errors:
            for error_msg, phones in whatsapp_errors.items():
                if len(phones) > 3:
                    messages.error(
                        request,
                        f'❌ WhatsApp Error ({len(phones)} recipients): {error_msg}'
                    )
                else:
                    for phone in phones[:3]:  # Limit to 3 to avoid spam
                        messages.error(request, f'❌ WhatsApp {phone}: {error_msg}')
        
        # Update notification statistics
        notification.email_sent_count = email_sent
        notification.email_failed_count = email_failed
        notification.sms_sent_count = sms_sent
        notification.sms_failed_count = sms_failed
        notification.whatsapp_sent_count = whatsapp_sent
        notification.whatsapp_failed_count = whatsapp_failed
        notification.mark_as_sent()
        
        # Store detailed error information in notes
        error_summary = []
        if email_errors:
            error_summary.append(f"Email Errors: {len(email_errors)} error type(s) affecting {sum(len(emails) for emails in email_errors.values())} recipient(s)")
        if sms_errors:
            error_summary.append(f"SMS Errors: {len(sms_errors)} error type(s) affecting {sum(len(phones) for phones in sms_errors.values())} recipient(s)")
        if whatsapp_errors:
            error_summary.append(f"WhatsApp Errors: {len(whatsapp_errors)} error type(s) affecting {sum(len(phones) for phones in whatsapp_errors.values())} recipient(s)")
        
        notes_update = f'\n[Manually sent on {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}]'
        if error_summary:
            notes_update += '\n' + '\n'.join(error_summary)
        notification.notes = (notification.notes or '') + notes_update
        notification.save()
        
        # Calculate totals and success rates
        total_recipients = len(email_subscribers) + len(sms_subscribers) + len(whatsapp_subscribers)
        total_sent = email_sent + sms_sent + whatsapp_sent
        total_failed = email_failed + sms_failed + whatsapp_failed
        success_rate = (total_sent / total_recipients * 100) if total_recipients > 0 else 0
        
        # Create detailed success message
        success_parts = [f'✅ Notification sent successfully!']
        success_parts.append(f'\n📊 Summary: {total_sent} sent, {total_failed} failed out of {total_recipients} total recipients ({success_rate:.1f}% success rate)')
        
        if notification.send_to_email:
            email_rate = (email_sent / len(email_subscribers) * 100) if email_subscribers else 0
            success_parts.append(f'\n📧 Email: {email_sent} sent, {email_failed} failed ({email_rate:.1f}% success)')
        
        if notification.send_to_sms:
            sms_rate = (sms_sent / len(sms_subscribers) * 100) if sms_subscribers else 0
            success_parts.append(f'\n📱 SMS: {sms_sent} sent, {sms_failed} failed ({sms_rate:.1f}% success)')
        
        if notification.send_to_whatsapp:
            whatsapp_rate = (whatsapp_sent / len(whatsapp_subscribers) * 100) if whatsapp_subscribers else 0
            success_parts.append(f'\n💬 WhatsApp: {whatsapp_sent} sent, {whatsapp_failed} failed ({whatsapp_rate:.1f}% success)')
        
        messages.success(request, ''.join(success_parts))
        
        # Store error details in session for display on detail page
        request.session[f'notification_{pk}_errors'] = {
            'email_errors': {k: list(v) for k, v in email_errors.items()},
            'sms_errors': {k: list(v) for k, v in sms_errors.items()},
            'whatsapp_errors': {k: list(v) for k, v in whatsapp_errors.items()},
        }
        
        return redirect('manage:notifications_detail', pk=pk)


class NotificationDeleteView(StaffRequiredMixin, DeleteView):
    """Delete a scheduled notification."""
    model = ScheduledNotification
    template_name = 'admin/notifications/delete.html'
    success_url = reverse_lazy('manage:notifications_list')
    
    def delete(self, request, *args, **kwargs):
        notification = self.get_object()
        messages.success(request, f'Notification "{notification.title}" has been deleted.')
        return super().delete(request, *args, **kwargs)


class CounselingBookingApproveView(StaffRequiredMixin, View):
    """Approve a counseling booking and send notifications."""
    def post(self, request, pk):
        booking = get_object_or_404(CounselingBooking, pk=pk)
        
        if booking.status != CounselingBooking.STATUS_PENDING:
            messages.error(request, 'Only pending bookings can be approved.')
            return redirect('manage:counseling_detail', pk=pk)
        
        # Update booking status
        booking.status = CounselingBooking.STATUS_APPROVED
        if not booking.approved_date:
            booking.approved_date = booking.preferred_date
        if not booking.approved_time:
            booking.approved_time = booking.preferred_time
        booking.save()
        
        # Send notifications (email and SMS)
        notification_result = send_booking_approval_notifications(booking)
        
        # Build detailed success/error message
        if notification_result['success']:
            success_parts = ['Booking approved and notifications sent successfully!']
            if notification_result['email_sent']:
                success_parts.append(f'✅ Email sent to {booking.email}')
            if notification_result['sms_sent']:
                success_parts.append(f'✅ SMS sent to {booking.phone}')
            if notification_result['admin_notification_sent']:
                success_parts.append('✅ Admin team notified')
            messages.success(request, ' | '.join(success_parts))
        else:
            # Some notifications failed - provide detailed error message
            error_parts = ['Booking approved, but some notifications failed:']
            
            if notification_result['errors']['email']:
                error_parts.append(f'❌ Email failed: {notification_result["errors"]["email"]}')
            elif booking.email and not notification_result['email_sent']:
                error_parts.append('❌ Email failed: Unknown error')
            
            if notification_result['errors']['sms']:
                error_parts.append(f'❌ SMS failed: {notification_result["errors"]["sms"]}')
            elif not notification_result['sms_sent']:
                error_parts.append('❌ SMS failed: Unknown error')
            
            if notification_result['errors']['admin']:
                error_parts.append(f'❌ Admin notification failed: {notification_result["errors"]["admin"]}')
            elif not notification_result['admin_notification_sent']:
                error_parts.append('❌ Admin notification failed: Unknown error')
            
            # Also show what succeeded
            if notification_result['email_sent']:
                error_parts.append(f'✅ Email sent to {booking.email}')
            if notification_result['sms_sent']:
                error_parts.append(f'✅ SMS sent to {booking.phone}')
            if notification_result['admin_notification_sent']:
                error_parts.append('✅ Admin team notified')
            
            messages.warning(request, ' | '.join(error_parts))
        
        # Create Google Calendar event
        try:
            create_google_calendar_event(booking)
            messages.success(request, 'Google Calendar event created successfully.')
        except Exception as e:
            messages.warning(request, f'Google Calendar event creation failed: {str(e)}')
        
        return redirect('manage:counseling_detail', pk=pk)


class CounselingBookingRejectView(StaffRequiredMixin, View):
    """Reject a counseling booking."""
    def post(self, request, pk):
        booking = get_object_or_404(CounselingBooking, pk=pk)
        
        if booking.status != CounselingBooking.STATUS_PENDING:
            messages.error(request, 'Only pending bookings can be rejected.')
            return redirect('manage:counseling_detail', pk=pk)
        
        booking.status = CounselingBooking.STATUS_REJECTED
        booking.save()
        
        messages.success(request, 'Booking rejected successfully.')
        return redirect('manage:counseling_detail', pk=pk)


# ==================== DEVOTION SERIES AJAX ====================

class DevotionSeriesCreateAjaxView(StaffRequiredMixin, View):
    """Create a new devotion series via AJAX."""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            title = data.get('title', '').strip()
            description = data.get('description', '').strip()
            start_date = data.get('start_date') or None
            end_date = data.get('end_date') or None
            
            if not title:
                return JsonResponse({'success': False, 'error': 'Title is required'}, status=400)
            
            # Check if series with this title already exists
            if DevotionSeries.objects.filter(title=title).exists():
                return JsonResponse({'success': False, 'error': 'A series with this title already exists'}, status=400)
            
            # Create the series
            series = DevotionSeries.objects.create(
                title=title,
                description=description,
                start_date=start_date,
                end_date=end_date,
                is_active=True
            )
            
            return JsonResponse({
                'success': True,
                'series': {
                    'id': series.id,
                    'title': series.title,
                    'slug': series.slug
                }
            })
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)